"""Export leads to CSV and Excel with tier-based ranking."""

import csv
import json
import os
import sys
from datetime import datetime

from collections import defaultdict

from config import (
    BUSINESS_TYPE_TIERS,
    CATEGORY_DISPLAY_NAMES,
)
from db import get_connection, migrate_db
from enrich import is_provider_name, normalize_phone

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "data")


def display_category(category):
    """Convert internal category to human-readable label."""
    return CATEGORY_DISPLAY_NAMES.get(category, category or "Unknown")


def compute_tier(lead):
    """
    Assign prospect tier A/B/C/D.

    A = "Call This Week": Right type (tier 1-2), 2+ distinct locations, 10+ min drive
    B = "Call This Month": Right type + expansion signals, or right type + strong reviews
    C = "Keep On List": Right type but no expansion signals
    D = "Skip": Hospital system, too close (<10 min), wrong type
    """
    category = lead["business_category"] or ""
    type_tier = BUSINESS_TYPE_TIERS.get(category, 3)
    distinct_locs = lead["distinct_location_count"] or 1
    drive_min = lead["drive_time_minutes"]
    rating = lead["rating"] or 0
    review_count = lead["rating_count"] or 0
    has_signals = bool(lead["multi_location_signals"])

    # Tier D: wrong business type
    if type_tier == 0:
        return "D"
    # Tier D: already a neighbor (strictly < 10 min; exactly 10.0 = included)
    if drive_min is not None and drive_min < 10:
        return "D"

    # Tier A: proven multi-location chain, right type, right distance
    if (type_tier in (1, 2) and distinct_locs >= 2
            and drive_min is not None and drive_min >= 10):
        return "A"

    # Tier B checks
    if type_tier in (1, 2):
        if has_signals or distinct_locs >= 2:
            return "B"
        if (drive_min is not None and drive_min >= 10
                and rating >= 4.0 and review_count >= 50):
            return "B"
    if type_tier == 3 and distinct_locs >= 2 and drive_min is not None and drive_min >= 10:
        return "B"

    return "C"


def drive_time_sort_score(minutes):
    """Score for sorting: 10-25 min sweet spot ranks highest."""
    if minutes is None:
        return 0
    if 10 <= minutes <= 25:
        return 2
    return 1


def pick_representative(group):
    """Pick the best representative row for an org group.

    Prefer: practice name > has email > 10-25 min drive > highest rating > most reviews.
    """
    def score(lead):
        return (
            0 if is_provider_name(lead["name"]) else 1,  # practice name preferred
            1 if lead["email"] else 0,
            drive_time_sort_score(lead["drive_time_minutes"]),
            lead["rating"] or 0,
            lead["rating_count"] or 0,
        )
    return max(group, key=score)


def get_all_leads(conn):
    """Fetch all businesses with org info, compute tier, return sorted list of dicts."""
    rows = conn.execute("""
        SELECT
            b.id, b.name, b.address, b.phone, b.website, b.email,
            b.description, b.rating, b.rating_count, b.primary_type,
            b.business_category, b.drive_time_minutes, b.drive_zone,
            b.distance_miles, b.search_query, b.multi_location_signals,
            b.organization_id,
            o.name as org_name, o.website_domain as org_domain,
            o.location_count, o.distinct_location_count
        FROM businesses b
        LEFT JOIN organizations o ON b.organization_id = o.id
    """).fetchall()

    leads = []
    for row in rows:
        lead = dict(row)
        lead["tier"] = compute_tier(lead)
        leads.append(lead)

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}
    leads.sort(key=lambda l: (
        tier_order.get(l["tier"], 4),
        -(l["distinct_location_count"] or 1),
        -drive_time_sort_score(l["drive_time_minutes"]),
        -(1 if l["email"] else 0),
        -(l["rating"] or 0),
        -(l["rating_count"] or 0),
    ))

    return leads


LEAD_CSV_HEADERS = [
    "Tier", "Name", "Category", "Phone", "Email", "Website", "Address",
    "Drive Time (min)", "Drive Zone", "Distinct Locations", "Org Name",
    "Rating", "Reviews", "Description", "Multi-Location Signals",
]


def lead_to_row(lead):
    """Convert a lead dict to a CSV row."""
    return [
        lead["tier"],
        lead["name"],
        display_category(lead["business_category"]),
        lead["phone"] or "",
        lead["email"] or "",
        lead["website"] or "",
        lead["address"] or "",
        lead["drive_time_minutes"] or "",
        lead["drive_zone"] or "",
        lead["distinct_location_count"] or 1,
        lead["org_name"] or "",
        lead["rating"] or "",
        lead["rating_count"] or "",
        lead["description"] or "",
        lead["multi_location_signals"] or "",
    ]


def export_ranked_csv(leads):
    """Export all leads ranked by tier to CSV."""
    filepath = os.path.join(OUTPUT_DIR, "leads_ranked.csv")
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(LEAD_CSV_HEADERS)
        for lead in leads:
            writer.writerow(lead_to_row(lead))
    print(f"CSV exported: {filepath} ({len(leads)} leads)")


MAX_DRIVE_TIME_PROSPECTS = 45


def export_top_prospects_csv(leads):
    """Export Tier A + B leads, deduplicated to one row per practice."""
    top = [l for l in leads if l["tier"] in ("A", "B")]

    # Filter out far-flung locations
    top = [l for l in top if l["drive_time_minutes"] is None
           or l["drive_time_minutes"] <= MAX_DRIVE_TIME_PROSPECTS]

    # Dedup by organization_id
    org_groups = defaultdict(list)
    for lead in top:
        key = lead.get("organization_id") or f"solo_{lead['id']}"
        org_groups[key].append(lead)

    deduped = []
    for group in org_groups.values():
        deduped.append(pick_representative(group))

    # Safety net: phone-number dedup for anything org detection missed
    phone_seen = {}
    final = []
    for lead in deduped:
        phone = normalize_phone(lead.get("phone"))
        if phone and phone in phone_seen:
            # Skip — already have a row for this phone number
            continue
        if phone:
            phone_seen[phone] = True
        final.append(lead)

    # Re-sort after dedup
    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}
    final.sort(key=lambda l: (
        tier_order.get(l["tier"], 4),
        -(l["distinct_location_count"] or 1),
        -drive_time_sort_score(l["drive_time_minutes"]),
        -(1 if l["email"] else 0),
        -(l["rating"] or 0),
        -(l["rating_count"] or 0),
    ))

    filepath = os.path.join(OUTPUT_DIR, "top_prospects.csv")
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(LEAD_CSV_HEADERS)
        for lead in final:
            writer.writerow(lead_to_row(lead))
    print(f"CSV exported: {filepath} ({len(final)} unique practices, "
          f"deduped from {len(top)} entries)")


def export_organizations_csv(conn):
    """Export multi-location organizations to CSV."""
    orgs = conn.execute("""
        SELECT
            o.id, o.name, o.website_domain, o.location_count,
            o.distinct_location_count,
            GROUP_CONCAT(DISTINCT b.business_category) as categories,
            GROUP_CONCAT(DISTINCT b.address) as addresses
        FROM organizations o
        JOIN businesses b ON b.organization_id = o.id
        WHERE o.distinct_location_count >= 2 OR o.location_count >= 2
        GROUP BY o.id
        ORDER BY o.distinct_location_count DESC, o.location_count DESC
    """).fetchall()

    filepath = os.path.join(OUTPUT_DIR, "organizations.csv")
    headers = ["Org Name", "Domain", "Distinct Locations", "Total Entries",
               "Category", "Addresses"]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for org in orgs:
            cats = (org["categories"] or "").split(",")
            writer.writerow([
                org["name"],
                org["website_domain"] or "",
                org["distinct_location_count"] or 1,
                org["location_count"] or 1,
                display_category(cats[0] if cats else ""),
                org["addresses"] or "",
            ])

    print(f"CSV exported: {filepath} ({len(orgs)} organizations)")


def export_excel(conn, leads):
    """Export to Excel with tier-based sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Skipping Excel export.")
        return

    filepath = os.path.join(OUTPUT_DIR, "leads.xlsx")
    wb = Workbook()

    headers = [
        "Rank", "Tier", "Name", "Category", "Phone", "Email", "Website",
        "Address", "Drive Time (min)", "Drive Zone", "Distinct Locations",
        "Org Name", "Rating", "Reviews", "Description",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    tier_fills = {
        "A": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "B": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "C": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "D": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }

    def write_lead_sheet(ws, sheet_leads):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for rank, lead in enumerate(sheet_leads, 1):
            row = rank + 1
            values = [
                rank, lead["tier"], lead["name"],
                display_category(lead["business_category"]),
                lead["phone"] or "", lead["email"] or "",
                lead["website"] or "", lead["address"] or "",
                lead["drive_time_minutes"], lead["drive_zone"] or "",
                lead["distinct_location_count"] or 1,
                lead["org_name"] or "", lead["rating"],
                lead["rating_count"], lead["description"] or "",
            ]
            fill = tier_fills.get(lead["tier"])
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if fill:
                    cell.fill = fill

        for col in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, min(len(sheet_leads) + 2, 100))
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"

    # Sheet 1: All Leads
    ws1 = wb.active
    ws1.title = "All Leads"
    write_lead_sheet(ws1, leads)

    # Sheet 2: Top Prospects (A + B)
    top = [l for l in leads if l["tier"] in ("A", "B")]
    ws2 = wb.create_sheet("Top Prospects")
    write_lead_sheet(ws2, top)

    # Sheet 3: Multi-Location Orgs
    ws3 = wb.create_sheet("Multi-Location Orgs")
    org_headers = ["Org Name", "Domain", "Distinct Locations", "Total Entries"]
    for col, h in enumerate(org_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    orgs = conn.execute("""
        SELECT name, website_domain, distinct_location_count, location_count
        FROM organizations
        WHERE distinct_location_count >= 2 OR location_count >= 2
        ORDER BY distinct_location_count DESC, location_count DESC
    """).fetchall()

    for i, org in enumerate(orgs, 2):
        ws3.cell(row=i, column=1, value=org["name"])
        ws3.cell(row=i, column=2, value=org["website_domain"] or "")
        ws3.cell(row=i, column=3, value=org["distinct_location_count"] or 1)
        ws3.cell(row=i, column=4, value=org["location_count"] or 1)
    ws3.freeze_panes = "A2"

    wb.save(filepath)
    print(f"Excel exported: {filepath}")


def print_summary(leads):
    """Print tier summary to stdout."""
    tier_counts = {"A": 0, "B": 0, "C": 0, "D": 0}
    for lead in leads:
        tier_counts[lead["tier"]] = tier_counts.get(lead["tier"], 0) + 1

    print(f"\nTier A (Call This Week): {tier_counts['A']} leads")
    print(f"Tier B (Call This Month): {tier_counts['B']} leads")
    print(f"Tier C (Keep On List): {tier_counts['C']} leads")
    print(f"Tier D (Skip): {tier_counts['D']} leads")

    if tier_counts["A"] < 5:
        print(f"\n  Note: Only {tier_counts['A']} Tier A leads — scoring may need calibration.")
    elif tier_counts["A"] > 200:
        print(f"\n  Note: {tier_counts['A']} Tier A leads — scoring may be too generous.")


def run_export():
    """Run the full export pipeline."""
    migrate_db()
    conn = get_connection()

    total = conn.execute("SELECT COUNT(*) FROM businesses").fetchone()[0]
    if total == 0:
        print("No businesses in database. Run collect.py first.")
        conn.close()
        sys.exit(1)

    print(f"Exporting {total} businesses...\n")

    leads = get_all_leads(conn)
    export_ranked_csv(leads)
    export_top_prospects_csv(leads)
    export_organizations_csv(conn)
    export_excel(conn, leads)
    print_summary(leads)

    conn.close()


if __name__ == "__main__":
    run_export()
